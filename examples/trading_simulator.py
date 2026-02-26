"""
模拟交易系统 - 支持资金模式 & 净值模式
支持单品种/多品种仓位文件
"""

import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---------- 中文字体 ----------
plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False


# ============================================================
# 工具函数
# ============================================================

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill(hex_color="1F497D"):
    return PatternFill("solid", fgColor=hex_color)

def _apply_header(ws, row, cols, texts, widths=None, hex_color="1F497D"):
    """写表头"""
    for i, (col, text) in enumerate(zip(cols, texts)):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _header_fill(hex_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        if widths:
            ws.column_dimensions[get_column_letter(col)].width = widths[i]

def _style_data_row(ws, row, col_start, col_end, fmt_map=None):
    """美化数据行"""
    fill = PatternFill("solid", fgColor="EBF1DE") if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = _border()
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt_map and col in fmt_map:
            cell.number_format = fmt_map[col]


# ============================================================
# 评价指标计算
# ============================================================

def calc_metrics(nav_series: pd.Series, rf: float = 0.0, periods_per_year: int = 252) -> dict:
    """
    输入: nav_series，以1为基准的净值序列（或资金序列/初始资金）
    rf:   年化无风险利率
    """
    nav = nav_series.dropna()
    if len(nav) < 2:
        return {}

    daily_ret = nav.pct_change().dropna()
    total_ret = nav.iloc[-1] / nav.iloc[0] - 1
    n_days = len(nav) - 1
    annual_ret = (1 + total_ret) ** (periods_per_year / n_days) - 1

    # 最大回撤
    peak = nav.cummax()
    drawdown = (nav - peak) / peak
    max_dd = drawdown.min()
    max_dd_end = drawdown.idxmin()
    max_dd_start = nav[:max_dd_end].idxmax()

    # 夏普比率
    rf_daily = (1 + rf) ** (1 / periods_per_year) - 1
    excess = daily_ret - rf_daily
    sharpe = (excess.mean() / excess.std() * np.sqrt(periods_per_year)
              if excess.std() > 0 else np.nan)

    # 卡玛比率
    calmar = annual_ret / abs(max_dd) if max_dd != 0 else np.nan

    # 索提诺比率
    downside = daily_ret[daily_ret < rf_daily]
    downside_std = downside.std()
    sortino = (excess.mean() / downside_std * np.sqrt(periods_per_year)
               if downside_std > 0 else np.nan)

    # 年化波动率
    annual_vol = daily_ret.std() * np.sqrt(periods_per_year)

    return {
        "总收益率": total_ret,
        "年化收益率": annual_ret,
        "年化波动率": annual_vol,
        "夏普比率": sharpe,
        "索提诺比率": sortino,
        "最大回撤": max_dd,
        "最大回撤开始": max_dd_start,
        "最大回撤结束": max_dd_end,
        "卡玛比率": calmar,
        "交易天数": n_days,
    }


# ============================================================
# 核心：TradingSimulator
# ============================================================

class TradingSimulator:
    """
    模拟交易系统

    Parameters
    ----------
    position_df : pd.DataFrame
        列：datetime + 各标的代码（值 0~1，表示目标仓位比例）
    price_df : pd.DataFrame
        列：datetime, wind_code, OPEN, CLOSE, HIGH, LOW
    mode : str
        'capital' 资金模式 | 'nav' 净值模式
    initial_capital : float
        起始资金（资金模式有效）
    min_lot : int
        最低买入手数（每手股数），默认 100
    commission_rate : float
        双边手续费率
    stamp_duty : float
        卖出印花税率（债券ETF默认0）
    friction_cost : float
        额外摩擦成本率（双边）
    exec_timing : str
        'next_open' 下一日开盘价成交（默认） | 'same_close' 当日收盘价成交
    start_date : str | None
    end_date   : str | None
    rf : float
        年化无风险利率，用于夏普等指标
    """

    def __init__(
        self,
        position_df: pd.DataFrame,
        price_df: pd.DataFrame,
        mode: str = "capital",
        initial_capital: float = 1_000_000,
        min_lot: int = 100,
        commission_rate: float = 0.0003,
        stamp_duty: float = 0.0,
        friction_cost: float = 0.0001,
        exec_timing: str = "next_open",
        start_date=None,
        end_date=None,
        rf: float = 0.0,
    ):
        self.mode = mode
        self.initial_capital = initial_capital
        self.min_lot = min_lot
        self.commission_rate = commission_rate
        self.stamp_duty = stamp_duty
        self.friction_cost = friction_cost
        self.exec_timing = exec_timing
        self.rf = rf

        # ---------- 整理价格数据 ----------
        price = price_df.copy()
        price["datetime"] = pd.to_datetime(price["datetime"])
        price = price.set_index(["datetime", "wind_code"]).sort_index()
        self.price = price

        # ---------- 整理仓位数据 ----------
        pos = position_df.copy()
        pos["datetime"] = pd.to_datetime(pos["datetime"])
        pos = pos.set_index("datetime").sort_index()
        self.securities = [c for c in pos.columns]
        self.pos = pos

        # ---------- 日期范围 ----------
        all_trade_dates = sorted(
            price.index.get_level_values("datetime").unique()
        )
        if start_date:
            all_trade_dates = [d for d in all_trade_dates if d >= pd.to_datetime(start_date)]
        if end_date:
            all_trade_dates = [d for d in all_trade_dates if d <= pd.to_datetime(end_date)]
        self.trade_dates = all_trade_dates

        # ---------- 结果容器 ----------
        self.daily_records = []       # 每日快照
        self.trade_records = []       # 调仓记录
        self.metrics = {}
        self.nav_series = None

    # ----------------------------------------------------------
    # 内部辅助
    # ----------------------------------------------------------

    def _get_price(self, date, code, col="CLOSE"):
        try:
            return float(self.price.loc[(date, code), col])
        except KeyError:
            return np.nan

    def _get_pos(self, date, code):
        """获取仓位文件中某日某标的的目标仓位（0或1），不存在则返回上一有效值"""
        try:
            return float(self.pos.loc[date, code])
        except KeyError:
            # 找最近的历史仓位
            before = self.pos[self.pos.index <= date]
            if len(before) == 0:
                return 0.0
            return float(before[code].iloc[-1])

    def _calc_buy_cost(self, price, qty):
        """买入总成本（含手续费+摩擦）"""
        return qty * price * (1 + self.commission_rate + self.friction_cost)

    def _calc_sell_revenue(self, price, qty):
        """卖出净收入（扣除手续费+印花税+摩擦）"""
        return qty * price * (1 - self.commission_rate - self.stamp_duty - self.friction_cost)

    def _calc_commission(self, price, qty, direction):
        base = price * qty
        comm = base * self.commission_rate
        friction = base * self.friction_cost
        tax = base * self.stamp_duty if direction == "sell" else 0.0
        return comm, tax, friction

    # ----------------------------------------------------------
    # 资金模式
    # ----------------------------------------------------------

    def _run_capital(self):
        cash = self.initial_capital
        # 每个标的的持仓信息
        holdings = {code: {"shares": 0, "cost_price": 0.0} for code in self.securities}
        prev_target = {code: 0.0 for code in self.securities}  # 前一日目标仓位

        for i, date in enumerate(self.trade_dates):
            # --- 获取本日目标仓位 ---
            target = {code: self._get_pos(date, code) for code in self.securities}

            # --- 判断是否需要调仓 ---
            for code in self.securities:
                if target[code] == prev_target[code]:
                    continue

                # 确定执行价格
                if self.exec_timing == "next_open":
                    exec_price = self._get_price(date, code, "OPEN")
                else:
                    exec_price = self._get_price(date, code, "CLOSE")

                if np.isnan(exec_price) or exec_price <= 0:
                    continue

                # 买入
                if target[code] > prev_target[code]:
                    portfolio_value = cash + sum(
                        holdings[c]["shares"] * (self._get_price(date, c, "CLOSE") or 0)
                        for c in self.securities
                    )
                    # 按目标仓位比例分配资金（简化：此处仅支持0/1仓位）
                    alloc_cash = cash * target[code]  # 可用于买入的资金
                    max_amount = alloc_cash / (1 + self.commission_rate + self.friction_cost)
                    shares_raw = max_amount / exec_price
                    shares = int(shares_raw // self.min_lot) * self.min_lot
                    if shares <= 0:
                        continue
                    total_cost = self._calc_buy_cost(exec_price, shares)
                    comm, tax, fric = self._calc_commission(exec_price, shares, "buy")
                    cash -= total_cost
                    holdings[code]["shares"] = shares
                    holdings[code]["cost_price"] = exec_price

                    self.trade_records.append({
                        "调仓日期": date,
                        "标的代码": code,
                        "方向": "买入",
                        "执行价格": round(exec_price, 4),
                        "数量(股)": shares,
                        "成交金额": round(exec_price * shares, 2),
                        "手续费": round(comm, 2),
                        "印花税": round(tax, 2),
                        "摩擦成本": round(fric, 2),
                        "净现金流": round(-total_cost, 2),
                        "剩余现金": round(cash, 2),
                    })

                # 卖出
                else:
                    shares = holdings[code]["shares"]
                    if shares <= 0:
                        holdings[code]["shares"] = 0
                        continue
                    revenue = self._calc_sell_revenue(exec_price, shares)
                    comm, tax, fric = self._calc_commission(exec_price, shares, "sell")
                    cash += revenue

                    self.trade_records.append({
                        "调仓日期": date,
                        "标的代码": code,
                        "方向": "卖出",
                        "执行价格": round(exec_price, 4),
                        "数量(股)": shares,
                        "成交金额": round(exec_price * shares, 2),
                        "手续费": round(comm, 2),
                        "印花税": round(tax, 2),
                        "摩擦成本": round(fric, 2),
                        "净现金流": round(revenue, 2),
                        "剩余现金": round(cash, 2),
                    })
                    holdings[code]["shares"] = 0
                    holdings[code]["cost_price"] = 0.0

            # --- 计算当日总资产 ---
            position_value = 0.0
            for code in self.securities:
                close_p = self._get_price(date, code, "CLOSE")
                if not np.isnan(close_p):
                    position_value += holdings[code]["shares"] * close_p

            total_assets = cash + position_value

            # --- 记录每日快照 ---
            rec = {
                "日期": date,
                "总资产": round(total_assets, 2),
                "现金": round(cash, 2),
                "持仓市值": round(position_value, 2),
                "净值": round(total_assets / self.initial_capital, 6),
            }
            for code in self.securities:
                rec[f"{code}_目标仓位"] = target[code]
                rec[f"{code}_持仓股数"] = holdings[code]["shares"]
                close_p = self._get_price(date, code, "CLOSE")
                rec[f"{code}_收盘价"] = round(close_p, 4) if not np.isnan(close_p) else np.nan
                rec[f"{code}_持仓市值"] = round(
                    holdings[code]["shares"] * close_p if not np.isnan(close_p) else 0, 2
                )
            self.daily_records.append(rec)
            prev_target = target.copy()

        daily_df = pd.DataFrame(self.daily_records)
        daily_df["当日损益"] = daily_df["总资产"].diff().fillna(
            daily_df["总资产"].iloc[0] - self.initial_capital
        ).round(2)
        daily_df["累计损益"] = (daily_df["总资产"] - self.initial_capital).round(2)
        self.daily_df = daily_df
        self.trade_df = pd.DataFrame(self.trade_records) if self.trade_records else pd.DataFrame()
        self.nav_series = daily_df.set_index("日期")["净值"]

    # ----------------------------------------------------------
    # 净值模式
    # ----------------------------------------------------------

    def _run_nav(self):
        nav = 1.0
        prev_target = {code: 0.0 for code in self.securities}
        # 追踪每个标的的"入场净值"，用于计算贡献
        entry_price = {code: np.nan for code in self.securities}

        for i, date in enumerate(self.trade_dates):
            target = {code: self._get_pos(date, code) for code in self.securities}

            for code in self.securities:
                if target[code] == prev_target[code]:
                    continue
                # 买入
                if target[code] > prev_target[code]:
                    ep = self._get_price(date, code, "OPEN" if self.exec_timing == "next_open" else "CLOSE")
                    if not np.isnan(ep) and ep > 0:
                        entry_price[code] = ep
                # 卖出
                else:
                    ep = self._get_price(date, code, "OPEN" if self.exec_timing == "next_open" else "CLOSE")
                    if not np.isnan(ep) and not np.isnan(entry_price.get(code, np.nan)):
                        trade_ret = ep / entry_price[code] - 1
                        cost_rate = self.commission_rate + self.friction_cost  # buy side
                        sell_cost = self.commission_rate + self.stamp_duty + self.friction_cost
                        net_ret = (1 + trade_ret) * (1 - sell_cost) / (1 + cost_rate) - 1
                        nav = nav * (1 + net_ret * prev_target[code])

                    entry_price[code] = np.nan

                    sell_rec = {
                        "调仓日期": date,
                        "标的代码": code,
                        "方向": "卖出(净值)",
                        "执行价格": round(ep, 4) if not np.isnan(ep) else np.nan,
                        "净值": round(nav, 6),
                    }
                    self.trade_records.append(sell_rec)

            # --- 当日收盘净值（持仓的浮动盈亏） ---
            daily_nav = nav
            for code in self.securities:
                if target[code] > 0 and not np.isnan(entry_price.get(code, np.nan)):
                    close_p = self._get_price(date, code, "CLOSE")
                    ep = entry_price[code]
                    if not np.isnan(close_p) and not np.isnan(ep) and ep > 0:
                        # 买入成本折算（扣手续费）
                        actual_ep = ep * (1 + self.commission_rate + self.friction_cost)
                        daily_nav = nav * (1 + (close_p / actual_ep - 1) * target[code])

            rec = {"日期": date, "净值": round(daily_nav, 6)}
            for code in self.securities:
                rec[f"{code}_目标仓位"] = target[code]
                close_p = self._get_price(date, code, "CLOSE")
                rec[f"{code}_收盘价"] = round(close_p, 4) if not np.isnan(close_p) else np.nan
            self.daily_records.append(rec)
            prev_target = target.copy()

        daily_df = pd.DataFrame(self.daily_records)
        daily_df["当日涨跌"] = daily_df["净值"].pct_change().fillna(0).round(6)
        self.daily_df = daily_df
        self.trade_df = pd.DataFrame(self.trade_records) if self.trade_records else pd.DataFrame()
        self.nav_series = daily_df.set_index("日期")["净值"]

    # ----------------------------------------------------------
    # 主入口
    # ----------------------------------------------------------

    def run(self):
        if self.mode == "capital":
            self._run_capital()
        else:
            self._run_nav()
        self.metrics = calc_metrics(self.nav_series, rf=self.rf)
        return self

    # ----------------------------------------------------------
    # 打印指标
    # ----------------------------------------------------------

    def print_metrics(self):
        m = self.metrics
        if not m:
            print("请先调用 run()")
            return
        print("\n" + "=" * 50)
        print(f"  模式: {'资金模式' if self.mode == 'capital' else '净值模式'}")
        print("=" * 50)
        print(f"  总收益率    : {m.get('总收益率', 0):.2%}")
        print(f"  年化收益率  : {m.get('年化收益率', 0):.2%}")
        print(f"  年化波动率  : {m.get('年化波动率', 0):.2%}")
        print(f"  夏普比率    : {m.get('夏普比率', float('nan')):.4f}")
        print(f"  索提诺比率  : {m.get('索提诺比率', float('nan')):.4f}")
        print(f"  最大回撤    : {m.get('最大回撤', 0):.2%}")
        print(f"  最大回撤区间: {m.get('最大回撤开始', '')} ~ {m.get('最大回撤结束', '')}")
        print(f"  卡玛比率    : {m.get('卡玛比率', float('nan')):.4f}")
        print(f"  交易天数    : {m.get('交易天数', 0)}")
        if self.mode == "capital":
            last_assets = self.daily_df["总资产"].iloc[-1]
            print(f"  起始资金    : {self.initial_capital:,.2f}")
            print(f"  期末资产    : {last_assets:,.2f}")
            print(f"  绝对收益    : {last_assets - self.initial_capital:,.2f}")
        print("=" * 50)

    # ----------------------------------------------------------
    # 绘图
    # ----------------------------------------------------------

    def plot(self, save_path=None, show=True, benchmark_nav: pd.Series = None):
        fig = plt.figure(figsize=(16, 10))
        from matplotlib.gridspec import GridSpec
        gs = GridSpec(3, 2, figure=fig, hspace=0.4, wspace=0.35)

        ax_nav    = fig.add_subplot(gs[0, :])   # 净值/资金曲线（全宽）
        ax_dd     = fig.add_subplot(gs[1, :])   # 回撤曲线（全宽）
        ax_pos    = fig.add_subplot(gs[2, 0])   # 仓位图
        ax_metric = fig.add_subplot(gs[2, 1])   # 指标表格

        nav = self.nav_series
        dates = nav.index

        # --- 净值曲线 ---
        ax_nav.plot(dates, nav.values, color="#2878B5", lw=1.5, label="策略净值")
        if benchmark_nav is not None:
            bm = benchmark_nav.reindex(dates).ffill()
            ax_nav.plot(dates, bm.values, color="#F28522", lw=1.2, ls="--", label="基准净值")
        ax_nav.set_title("净值曲线" if self.mode == "nav" else "资金净值曲线", fontsize=12, fontweight="bold")
        ax_nav.set_ylabel("净值")
        ax_nav.legend(loc="upper left")
        ax_nav.grid(True, alpha=0.3)
        ax_nav.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
        ax_nav.xaxis.set_major_locator(mdates.YearLocator())
        fig.autofmt_xdate()

        # --- 回撤曲线 ---
        peak = nav.cummax()
        drawdown = (nav - peak) / peak * 100
        ax_dd.fill_between(dates, drawdown.values, 0, color="#D73027", alpha=0.5, label="回撤%")
        ax_dd.set_title("回撤曲线", fontsize=12, fontweight="bold")
        ax_dd.set_ylabel("回撤 (%)")
        ax_dd.legend(loc="lower left")
        ax_dd.grid(True, alpha=0.3)
        ax_dd.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
        ax_dd.xaxis.set_major_locator(mdates.YearLocator())

        # --- 仓位图 ---
        for code in self.securities:
            col = f"{code}_目标仓位"
            if col in self.daily_df.columns:
                pos_vals = self.daily_df.set_index("日期")[col]
                ax_pos.fill_between(pos_vals.index, pos_vals.values, alpha=0.5, label=code)
        ax_pos.set_title("仓位变化", fontsize=11, fontweight="bold")
        ax_pos.set_ylabel("仓位")
        ax_pos.set_ylim(-0.05, 1.15)
        ax_pos.legend(fontsize=8)
        ax_pos.grid(True, alpha=0.3)
        ax_pos.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
        ax_pos.xaxis.set_major_locator(mdates.YearLocator(2))

        # --- 指标表 ---
        ax_metric.axis("off")
        m = self.metrics
        rows = [
            ["总收益率",   f"{m.get('总收益率',0):.2%}"],
            ["年化收益率", f"{m.get('年化收益率',0):.2%}"],
            ["年化波动率", f"{m.get('年化波动率',0):.2%}"],
            ["夏普比率",   f"{m.get('夏普比率', float('nan')):.3f}"],
            ["索提诺比率", f"{m.get('索提诺比率', float('nan')):.3f}"],
            ["最大回撤",   f"{m.get('最大回撤',0):.2%}"],
            ["卡玛比率",   f"{m.get('卡玛比率', float('nan')):.3f}"],
            ["交易天数",   str(m.get('交易天数', 0))],
        ]
        tbl = ax_metric.table(
            cellText=rows,
            colLabels=["指标", "值"],
            cellLoc="center",
            loc="center",
            bbox=[0, 0, 1, 1],
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(10)
        for (r, c), cell in tbl.get_celld().items():
            if r == 0:
                cell.set_facecolor("#1F497D")
                cell.set_text_props(color="white", fontweight="bold")
            elif r % 2 == 0:
                cell.set_facecolor("#EBF1DE")
        ax_metric.set_title("评价指标", fontsize=11, fontweight="bold")

        fig.suptitle(
            f"模拟交易系统 - {'资金模式' if self.mode == 'capital' else '净值模式'} | "
            f"{str(dates[0])[:10]} ~ {str(dates[-1])[:10]}",
            fontsize=13, fontweight="bold", y=1.01
        )

        if save_path:
            fig.savefig(save_path, dpi=150, bbox_inches="tight")
            print(f"图表已保存: {save_path}")
        if show:
            plt.show()
        plt.close(fig)

    # ----------------------------------------------------------
    # 导出 Excel
    # ----------------------------------------------------------

    def to_excel(self, path: str):
        wb = Workbook()

        # ---- Sheet1: 仓位明细 ----
        ws1 = wb.active
        ws1.title = "仓位明细"
        self._write_position_sheet(ws1)

        # ---- Sheet2: 调仓记录 ----
        ws2 = wb.create_sheet("调仓记录")
        self._write_trade_sheet(ws2)

        # ---- Sheet3: 评价指标 ----
        ws3 = wb.create_sheet("评价指标")
        self._write_metrics_sheet(ws3)

        wb.save(path)
        print(f"Excel 已保存: {path}")

    def _write_position_sheet(self, ws):
        df = self.daily_df.copy()
        df["日期"] = df["日期"].astype(str)

        cols_order = ["日期", "净值"]
        if self.mode == "capital":
            cols_order += ["总资产", "现金", "持仓市值", "当日损益", "累计损益"]
        else:
            cols_order += ["当日涨跌"]

        for code in self.securities:
            for suffix in ["_目标仓位", "_持仓股数", "_收盘价", "_持仓市值"]:
                col = f"{code}{suffix}"
                if col in df.columns:
                    cols_order.append(col)

        # 仅保留存在的列
        cols_order = [c for c in cols_order if c in df.columns]
        df = df[cols_order]

        # 写表头
        headers = []
        rename_map = {
            "净值": "净值",
            "总资产": "总资产(元)",
            "现金": "现金(元)",
            "持仓市值": "持仓市值(元)",
            "当日损益": "当日损益(元)",
            "累计损益": "累计损益(元)",
            "当日涨跌": "当日涨跌",
        }
        for c in cols_order:
            headers.append(rename_map.get(c, c))

        _apply_header(ws, 1, range(1, len(headers) + 1), headers,
                      widths=[14] + [14] * (len(headers) - 1))

        # 格式映射
        fmt_map = {}
        for i, c in enumerate(cols_order, 1):
            if "仓位" in c or "涨跌" in c:
                fmt_map[i] = "0.00%"
            elif "价格" in c or "收盘价" in c or "净值" in c:
                fmt_map[i] = "0.0000"
            elif "元" in headers[i - 1] or "资产" in c or "损益" in c or "现金" in c or "市值" in c:
                fmt_map[i] = "#,##0.00"
            elif "股数" in c:
                fmt_map[i] = "#,##0"

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_row(ws, row_idx, 1, len(cols_order), fmt_map)

        ws.freeze_panes = "B2"

    def _write_trade_sheet(self, ws):
        if self.trade_df.empty:
            ws.cell(1, 1, "暂无调仓记录")
            return

        df = self.trade_df.copy()
        if "调仓日期" in df.columns:
            df["调仓日期"] = df["调仓日期"].astype(str)

        headers = df.columns.tolist()
        _apply_header(ws, 1, range(1, len(headers) + 1), headers,
                      widths=[14] * len(headers), hex_color="375623")

        fmt_map = {}
        for i, h in enumerate(headers, 1):
            if "价格" in h or "净值" in h:
                fmt_map[i] = "0.0000"
            elif any(k in h for k in ["金额", "手续费", "印花税", "摩擦", "现金流", "现金"]):
                fmt_map[i] = "#,##0.00"
            elif "数量" in h:
                fmt_map[i] = "#,##0"

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            _style_data_row(ws, row_idx, 1, len(headers), fmt_map)

        ws.freeze_panes = "B2"

    def _write_metrics_sheet(self, ws):
        m = self.metrics
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20

        title_rows = [
            ("模拟交易系统 - 评价指标", None),
            (f"模式: {'资金模式' if self.mode=='capital' else '净值模式'}", None),
            ("", None),
            ("指标", "值"),
        ]
        for r, (k, v) in enumerate(title_rows, 1):
            ws.cell(r, 1, k)
            if v is not None:
                ws.cell(r, 2, v)

        _apply_header(ws, 4, [1, 2], ["指标", "值"], hex_color="1F497D")

        data_rows = [
            ("总收益率",   m.get("总收益率", 0)),
            ("年化收益率", m.get("年化收益率", 0)),
            ("年化波动率", m.get("年化波动率", 0)),
            ("夏普比率",   m.get("夏普比率", float("nan"))),
            ("索提诺比率", m.get("索提诺比率", float("nan"))),
            ("最大回撤",   m.get("最大回撤", 0)),
            ("最大回撤开始", str(m.get("最大回撤开始", ""))),
            ("最大回撤结束", str(m.get("最大回撤结束", ""))),
            ("卡玛比率",   m.get("卡玛比率", float("nan"))),
            ("交易天数",   m.get("交易天数", 0)),
        ]
        if self.mode == "capital":
            last = self.daily_df["总资产"].iloc[-1]
            data_rows += [
                ("起始资金",   self.initial_capital),
                ("期末资产",   last),
                ("绝对收益",   last - self.initial_capital),
                ("调仓次数",   len(self.trade_df) if not self.trade_df.empty else 0),
            ]

        pct_keys = {"总收益率", "年化收益率", "年化波动率", "最大回撤"}
        money_keys = {"起始资金", "期末资产", "绝对收益"}

        for i, (key, val) in enumerate(data_rows, start=5):
            c1 = ws.cell(i, 1, key)
            c2 = ws.cell(i, 2, val)
            c1.border = _border()
            c2.border = _border()
            c1.alignment = Alignment(horizontal="center")
            c2.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c1.fill = PatternFill("solid", fgColor="EBF1DE")
                c2.fill = PatternFill("solid", fgColor="EBF1DE")
            if key in pct_keys:
                c2.number_format = "0.00%"
            elif key in money_keys:
                c2.number_format = "#,##0.00"
            elif key in {"夏普比率", "索提诺比率", "卡玛比率"}:
                c2.number_format = "0.0000"


# ============================================================
# 便捷函数
# ============================================================

def run_simulation(
    position_csv: str,
    price_csv: str,
    mode: str = "capital",
    output_dir: str = ".",
    initial_capital: float = 1_000_000,
    min_lot: int = 100,
    commission_rate: float = 0.0003,
    stamp_duty: float = 0.0,
    friction_cost: float = 0.0001,
    exec_timing: str = "next_open",
    start_date=None,
    end_date=None,
    rf: float = 0.0,
    show_plot: bool = True,
):
    print(f"\n加载数据...")
    pos_df = pd.read_csv(position_csv)
    price_df = pd.read_csv(price_csv)

    print(f"仓位文件: {pos_df.shape}, 价格文件: {price_df.shape}")

    sim = TradingSimulator(
        position_df=pos_df,
        price_df=price_df,
        mode=mode,
        initial_capital=initial_capital,
        min_lot=min_lot,
        commission_rate=commission_rate,
        stamp_duty=stamp_duty,
        friction_cost=friction_cost,
        exec_timing=exec_timing,
        start_date=start_date,
        end_date=end_date,
        rf=rf,
    )

    print(f"运行模拟 ({mode} 模式)...")
    sim.run()
    sim.print_metrics()

    os.makedirs(output_dir, exist_ok=True)
    prefix = f"{output_dir}/{mode}"

    # 图表
    sim.plot(save_path=f"{prefix}_chart.png", show=show_plot)

    # Excel
    sim.to_excel(f"{prefix}_result.xlsx")

    return sim
