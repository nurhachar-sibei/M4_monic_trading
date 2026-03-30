"""
excel_writer.py - Excel 输出模块

生成包含以下 Sheet 的 Excel 文件：
  1. 仓位明细一览 - 每日资产/净值快照
  2. 调仓记录     - 每次买卖记录
  3. 评价指标     - 综合绩效表
  4. 逐年收益     - 年度统计

另可生成"每日仓位明细"文件夹，每个交易日一个 Excel，
包含当日各持仓标的的 12 个详细字段。
"""

from __future__ import annotations

import os
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from .engine import SimulationResult
from .metrics import MetricsCalculator


# --------------------------------------------------------------------------- #
# 样式常量
# --------------------------------------------------------------------------- #

_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FILLS: Dict[str, str] = {
    "capital": "1F497D",   # 深蓝
    "trade":   "375623",   # 深绿
    "metrics": "7030A0",   # 紫色
    "yearly":  "C55A11",   # 橙色
}
_ROW_EVEN = PatternFill("solid", fgColor="EBF1DE")
_ROW_ODD  = PatternFill("solid", fgColor="FFFFFF")


def _hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_header(ws, row: int, headers: List[str], widths: List[float], color_key: str) -> None:
    fill = _hfill(_HEADER_FILLS.get(color_key, "1F497D"))
    for col, (text, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width


def _style_row(ws, row: int, n_cols: int, fmt_map: Optional[Dict[int, str]] = None) -> None:
    fill = _ROW_EVEN if row % 2 == 0 else _ROW_ODD
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt_map and col in fmt_map:
            cell.number_format = fmt_map[col]


# --------------------------------------------------------------------------- #
# 主写入类
# --------------------------------------------------------------------------- #

class ExcelWriter:
    """
    Excel 写入器

    Parameters
    ----------
    result : SimulationResult
    """

    def __init__(self, result: SimulationResult) -> None:
        self.result = result
        self._calc = MetricsCalculator(
            result.nav_series,
            rf=result.config.metrics.risk_free_rate if result.config else 0.0,
            periods_per_year=result.config.metrics.periods_per_year if result.config else 252,
        )
        self._metrics = self._calc.calculate()

    def write(self, path: str) -> None:
        """写出到 Excel 文件"""
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        wb = Workbook()

        # Sheet 1: 仓位明细一览
        ws1 = wb.active
        ws1.title = "仓位明细一览"
        self._write_position_sheet(ws1)

        # Sheet 2: 调仓记录
        ws2 = wb.create_sheet("调仓记录")
        self._write_trade_sheet(ws2)

        # Sheet 3: 评价指标
        ws3 = wb.create_sheet("评价指标")
        self._write_metrics_sheet(ws3)

        # Sheet 4: 逐年收益
        ws4 = wb.create_sheet("逐年收益")
        self._write_yearly_sheet(ws4)

        # Sheet 5: 净值序列（如果配置了bench）
        if self.result.bench_nav_series is not None:
            ws5 = wb.create_sheet("净值序列")
            self._write_nav_series_sheet(ws5)
            
            # Sheet 6: 每日收益率
            ws6 = wb.create_sheet("每日收益率")
            self._write_daily_returns_sheet(ws6)
            
            # Sheet 7: 月度收益
            ws7 = wb.create_sheet("月度收益")
            self._write_monthly_returns_sheet(ws7)

        wb.save(path)
        print(f"  Excel 已保存 → {path}")

    # ------------------------------------------------------------------ #
    # Sheet 1 - 仓位明细
    # ------------------------------------------------------------------ #

    def _write_position_sheet(self, ws) -> None:
        df = self.result.daily_df.copy()
        is_capital = self.result.mode == "capital"

        # 确定列顺序
        base_cols = ["日期", "净值"]
        if is_capital:
            base_cols += ["总资产", "现金", "持仓市值", "当日损益", "累计损益"]
        else:
            base_cols += ["当日涨跌"]

        sec_cols = []
        for code in self.result.securities:
            for suffix in ["_目标仓位", "_实际仓位", "_持仓股数", "_持仓份额", "_收盘价", "_持仓市值"]:
                col = f"{code}{suffix}"
                if col in df.columns:
                    sec_cols.append(col)

        all_cols = [c for c in base_cols + sec_cols if c in df.columns]
        df = df[all_cols]

        # 表头映射
        rename = {
            "总资产":    "总资产(元)",
            "现金":      "现金(元)",
            "持仓市值":  "持仓市值(元)",
            "当日损益":  "当日损益(元)",
            "累计损益":  "累计损益(元)",
        }
        headers = [rename.get(c, c) for c in all_cols]
        widths = [14] + [14] * (len(headers) - 1)
        # 标题列适当加宽
        for i, h in enumerate(headers):
            if "代码" in h or "日期" in h:
                widths[i] = 16

        ws.row_dimensions[1].height = 22
        _write_header(ws, 1, headers, widths, "capital")

        # 格式映射 (1-based column index)
        fmt_map: Dict[int, str] = {}
        for i, col in enumerate(all_cols, 1):
            h = headers[i - 1]
            if "仓位" in col or "涨跌" in col:
                fmt_map[i] = "0.00%"
            elif "净值" in col:
                fmt_map[i] = "0.000000"
            elif "收盘价" in col or "价格" in col:
                fmt_map[i] = "0.0000"
            elif "元" in h or "损益" in col or "资产" in col or "现金" in col or "市值" in col:
                fmt_map[i] = "#,##0.00"
            elif "股数" in col:
                fmt_map[i] = "#,##0"
            elif "份额" in col:
                fmt_map[i] = "0.00000000"

        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_data, start=1):
                v = str(val) if isinstance(val, pd.Timestamp) else val
                ws.cell(row=row_idx, column=col_idx, value=v)
            _style_row(ws, row_idx, len(all_cols), fmt_map)

        ws.freeze_panes = "B2"

    # ------------------------------------------------------------------ #
    # Sheet 2 - 调仓记录
    # ------------------------------------------------------------------ #

    def _write_trade_sheet(self, ws) -> None:
        df = self.result.trade_df.copy()
        if df.empty:
            ws.cell(1, 1, "暂无调仓记录")
            return

        if "调仓日期" in df.columns:
            df["调仓日期"] = df["调仓日期"].astype(str)

        headers = list(df.columns)
        widths = []
        for h in headers:
            if "日期" in h:
                widths.append(14)
            elif "代码" in h:
                widths.append(14)
            elif "方向" in h:
                widths.append(8)
            elif "金额" in h or "现金" in h:
                widths.append(16)
            else:
                widths.append(13)

        ws.row_dimensions[1].height = 22
        _write_header(ws, 1, headers, widths, "trade")

        fmt_map: Dict[int, str] = {}
        for i, h in enumerate(headers, 1):
            if "价格" in h or "净值" in h:
                fmt_map[i] = "0.0000"
            elif "仓位" in h:
                fmt_map[i] = "0.00%"
            elif any(k in h for k in ["金额", "手续费", "印花税", "摩擦", "现金流", "现金"]):
                fmt_map[i] = "#,##0.00"
            elif "数量" in h or "股" in h:
                fmt_map[i] = "#,##0"

        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            _style_row(ws, row_idx, len(headers), fmt_map)

        ws.freeze_panes = "B2"

    # ------------------------------------------------------------------ #
    # Sheet 3 - 评价指标
    # ------------------------------------------------------------------ #

    def _write_metrics_sheet(self, ws) -> None:
        m = self._metrics
        has_bench = self.result.bench_nav_series is not None
        
        # 如果有bench，使用三列表头：指标、策略、基准
        if has_bench:
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18
        else:
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 22

        mode_cn = "资金模式" if self.result.mode == "capital" else "净值模式"

        # 标题行
        title_cell = ws.cell(1, 1, f"模拟交易系统绩效报告 [{mode_cn}]")
        title_cell.font = Font(bold=True, size=13, color="1F497D")
        title_cell.alignment = Alignment(horizontal="left")

        nav = self.result.nav_series
        ws.cell(2, 1, f"回测区间：{str(nav.index[0])[:10]} ~ {str(nav.index[-1])[:10]}")

        # 表头
        if has_bench:
            _write_header(ws, 4, ["指标", "策略", "基准"], [22, 18, 18], "metrics")
            bench_m = self.result.bench_metrics
        else:
            _write_header(ws, 4, ["指标", "值"], [20, 22], "metrics")
            bench_m = {}

        rows_data = [
            ("--- 收益 ---", None),
            ("总收益率",      m.get("总收益率",   np.nan), bench_m.get("总收益率", np.nan) if has_bench else None),
            ("年化收益率",    m.get("年化收益率", np.nan), bench_m.get("年化收益率", np.nan) if has_bench else None),
            ("--- 风险 ---", None),
            ("年化波动率",    m.get("年化波动率", np.nan), bench_m.get("年化波动率", np.nan) if has_bench else None),
            ("最大回撤",      m.get("最大回撤",   np.nan), bench_m.get("最大回撤", np.nan) if has_bench else None),
            ("最大回撤开始",  str(m.get("最大回撤开始", ""))[:10], str(bench_m.get("最大回撤开始", ""))[:10] if has_bench else None),
            ("最大回撤结束",  str(m.get("最大回撤结束", ""))[:10], str(bench_m.get("最大回撤结束", ""))[:10] if has_bench else None),
            ("水下天数",      m.get("水下天数",   np.nan), bench_m.get("水下天数", np.nan) if has_bench else None),
            ("--- 风险调整 ---", None),
            ("夏普比率",      m.get("夏普比率",   np.nan), bench_m.get("夏普比率", np.nan) if has_bench else None),
            ("索提诺比率",    m.get("索提诺比率", np.nan), bench_m.get("索提诺比率", np.nan) if has_bench else None),
            ("卡玛比率",      m.get("卡玛比率",   np.nan), bench_m.get("卡玛比率", np.nan) if has_bench else None),
            ("--- 胜率 ---",  None),
            ("月度胜率",      m.get("月度胜率",   np.nan), bench_m.get("月度胜率", np.nan) if has_bench else None),
            ("年度胜率",      m.get("年度胜率",   np.nan), bench_m.get("年度胜率", np.nan) if has_bench else None),
        ]
        
        # 添加超额收益指标（如果有bench）
        if has_bench:
            ex = self.result.excess_metrics
            rows_data += [
                ("--- 超额收益 (策略-基准) ---", None),
                ("年化超额收益", ex.get("年化超额收益", np.nan), None),
                ("总超额收益", ex.get("总超额收益", np.nan), None),
                ("年化超额标准差", ex.get("年化超额标准差", np.nan), None),
                ("信息比率", ex.get("信息比率", np.nan), None),
                ("超额收益最大回撤", ex.get("超额收益最大回撤", np.nan), None),
                ("日度胜率", ex.get("日度胜率", np.nan), None),
                ("月度胜率", ex.get("月度胜率", np.nan), None),
                ("超额赔率", ex.get("超额赔率", np.nan), None),
            ]
        
        rows_data += [
            ("--- 基本信息 ---", None),
            ("交易天数",      m.get("交易天数",   0), bench_m.get("交易天数", 0) if has_bench else None),
            ("调仓次数",      self.result.n_trades, self.result.n_trades if has_bench else None),  # bench无调仓概念
        ]
        
        if self.result.mode == "capital" and self.result.config:
            ic = self.result.config.capital.initial_capital
            last_nav = float(self.result.nav_series.iloc[-1])
            if has_bench:
                bench_last_nav = float(self.result.bench_nav_series.iloc[-1])
                rows_data += [
                    ("--- 资金 ---",  None),
                    ("起始资金",      ic, ic),
                    ("期末资产",      ic * last_nav, ic * bench_last_nav),
                    ("绝对收益",      ic * (last_nav - 1.0), ic * (bench_last_nav - 1.0)),
                ]
            else:
                rows_data += [
                    ("--- 资金 ---",  None),
                    ("起始资金",      ic, None),
                    ("期末资产",      ic * last_nav, None),
                    ("绝对收益",      ic * (last_nav - 1.0), None),
                ]

        pct_keys = {"总收益率", "年化收益率", "年化波动率", "最大回撤", "月度胜率", "年度胜率", 
                    "年化超额收益", "总超额收益", "年化超额标准差", "超额收益最大回撤", 
                    "日度胜率", "月度胜率"}
        payoff_keys = {"超额赔率"}  # 赔率格式化为x.xx倍
        money_keys = {"起始资金", "期末资产", "绝对收益"}
        float4_keys = {"夏普比率", "索提诺比率", "卡玛比率", "信息比率"}

        section_fill = PatternFill("solid", fgColor="D9E1F2")

        for i, row in enumerate(rows_data, start=5):
            key = row[0]
            val_strategy = row[1]
            val_bench = row[2] if len(row) > 2 else None
            
            c1 = ws.cell(i, 1, key)
            c2 = ws.cell(i, 2, val_strategy)
            if has_bench and val_bench is not None:
                c3 = ws.cell(i, 3, val_bench)
                cells = (c1, c2, c3)
            else:
                cells = (c1, c2)
            
            for cell in cells:
                cell.border = _BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")

            is_section = key.startswith("---")
            if is_section:
                for cell in cells:
                    cell.fill = section_fill
                    cell.font = Font(bold=True, color="1F497D", italic=True)
            else:
                fill = _ROW_EVEN if i % 2 == 0 else _ROW_ODD
                for cell in cells:
                    cell.fill = fill
                # 数字格式
                for val, cell in [(val_strategy, c2), (val_bench, c3 if has_bench and val_bench is not None else None)]:
                    if cell is None:
                        continue
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        if key in pct_keys:
                            cell.number_format = "0.00%"
                        elif key in money_keys:
                            cell.number_format = "#,##0.00"
                        elif key in float4_keys:
                            cell.number_format = "0.0000"
                        elif key in payoff_keys:
                            cell.number_format = "0.00\"倍\""
                        elif key in {"交易天数", "水下天数", "调仓次数"}:
                            cell.number_format = "#,##0"

        ws.freeze_panes = "A5"

    # ------------------------------------------------------------------ #
    # Sheet 4 - 逐年收益
    # ------------------------------------------------------------------ #

    def _write_yearly_sheet(self, ws) -> None:
        yearly_df = self._calc.yearly_stats()
        if yearly_df.empty:
            ws.cell(1, 1, "数据不足")
            return

        yearly_df = yearly_df.reset_index()
        yearly_df.columns = ["年份", "年度收益率", "年内最大回撤"]

        headers = list(yearly_df.columns)
        widths = [10, 14, 16]

        ws.row_dimensions[1].height = 22
        _write_header(ws, 1, headers, widths, "yearly")

        fmt_map = {2: "0.00%", 3: "0.00%"}
        for row_idx, row_data in enumerate(yearly_df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            _style_row(ws, row_idx, len(headers), fmt_map)

        # 年度收益率色阶条件格式
        last_row = len(yearly_df) + 1
        ws.conditional_formatting.add(
            f"B2:B{last_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="num", mid_value=0, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )
        ws.conditional_formatting.add(
            f"C2:C{last_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                end_type="max", end_color="F8696B",
            ),
        )

    # ------------------------------------------------------------------ #
    # 每日仓位明细文件夹
    # ------------------------------------------------------------------ #

    def write_daily_folder(self, output_dir: str) -> None:
        """
        在 output_dir/每日仓位明细/ 下为每个交易日生成一个 Excel 文件。

        文件名格式：YYYYMMDD.xlsx
        每文件包含当日各持仓标的的 12 个字段（一行一个标的）：
          日期 | 资产代码 | 买入金额 | 当前金额 | 昨日收盘价 | 今日开盘价
          今日收盘价 | 持仓份额 | 今日收盘市值 | 今日收盘占资金权重
          当日损益 | 持仓期间累计损益
        """
        folder = os.path.join(output_dir, "每日仓位明细")
        os.makedirs(folder, exist_ok=True)

        df         = self.result.daily_df.copy()
        securities = self.result.securities
        is_capital = self.result.mode == "capital"

        HEADERS = [
            "日期", "资产代码", "买入价格", "当前价格",
            "买入金额", "当前金额",
            "昨日收盘价", "今日开盘价", "今日收盘价",
            "持仓份额", "今日收盘市值", "今日收盘占资金权重",
            "当日损益", "持仓期间累计损益",
        ]
        WIDTHS = [13, 14, 12, 12, 15, 15, 12, 12, 12, 12, 15, 18, 14, 18]

        # 资金模式：金额用 #,##0.00；净值模式：损益用 0.000000
        if is_capital:
            FMT_MAP = {
                3: "0.0000",     # 买入价格
                4: "0.0000",     # 当前价格
                5: "#,##0.00",   # 买入金额
                6: "#,##0.00",   # 当前金额
                7: "0.0000",     # 昨日收盘价
                8: "0.0000",     # 今日开盘价
                9: "0.0000",     # 今日收盘价
                10: "#,##0",     # 持仓份额（股数）
                11: "#,##0.00",  # 今日收盘市值
                12: "0.00%",     # 今日收盘占资金权重
                13: "#,##0.00",  # 当日损益
                14: "#,##0.00",  # 持仓期间累计损益
            }
        else:
            FMT_MAP = {
                3: "0.0000",      # 买入价格（加权平均执行价，元）
                4: "0.0000",      # 当前价格（收盘价，元）
                5: "0.000000",    # 买入金额（组合单位）
                6: "0.000000",    # 当前金额（组合单位）
                7: "0.0000",      # 昨日收盘价
                8: "0.0000",      # 今日开盘价
                9: "0.0000",      # 今日收盘价
                10: "0.00000000", # 持仓份额（小数份额）
                11: "0.000000",   # 今日收盘市值（组合单位）
                12: "0.00%",      # 今日收盘占资金权重
                13: "0.000000",   # 当日损益（组合单位）
                14: "0.000000",   # 持仓期间累计损益（组合单位）
            }

        n_days = len(df)
        for idx, row in df.iterrows():
            date     = row["日期"]
            date_str = str(date)[:10].replace("-", "")

            # 计算当日总资产（资金模式）
            total_assets = row.get("总资产", np.nan) if is_capital else np.nan
            nav          = row.get("净值", 1.0)

            rows_data = []
            for code in securities:
                close_p    = row.get(f"{code}_收盘价", np.nan)
                open_p     = row.get(f"{code}_开盘价", np.nan)
                prev_close = row.get(f"{code}_昨收价", np.nan)
                cost       = row.get(f"{code}_买入成本", np.nan)
                daily_pnl  = row.get(f"{code}_当日损益", np.nan)
                cum_pnl    = row.get(f"{code}_累计损益", np.nan)

                if is_capital:
                    shares  = int(row.get(f"{code}_持仓股数", 0))
                    mv      = row.get(f"{code}_持仓市值", 0.0)
                    weight  = (mv / total_assets
                               if (not np.isnan(total_assets) and total_assets > 0)
                               else np.nan)
                    # 买入价格 = 累计买入成本 ÷ 持仓股数（持仓为 0 时显示 nan）
                    buy_price = (cost / shares) if (shares > 0 and not np.isnan(cost)) else np.nan
                    # 当前价格 = 今日收盘价
                    cur_price = close_p
                else:
                    # 净值模式：持仓份额 = 实际小数份额，价格单位为元
                    actual_shares = row.get(f"{code}_持仓份额", 0.0)
                    weight        = row.get(f"{code}_实际仓位", 0.0)    # 日末真实权重
                    avg_ep        = row.get(f"{code}_平均买价", np.nan)  # 加权平均买价（元）
                    mv_nav        = (actual_shares * close_p
                                     if (actual_shares > 0 and not np.isnan(close_p))
                                     else 0.0)
                    # 买入价格 = 加权平均执行价（元）
                    buy_price = avg_ep
                    # 当前价格 = 今日收盘价（元）
                    cur_price = close_p
                    shares    = actual_shares

                rows_data.append([
                    str(date)[:10],                      # 日期
                    code,                                 # 资产代码
                    buy_price,                            # 买入价格
                    cur_price,                            # 当前价格
                    cost,                                 # 买入金额
                    mv if is_capital else mv_nav,         # 当前金额
                    prev_close,                           # 昨日收盘价
                    open_p,                               # 今日开盘价
                    close_p,                              # 今日收盘价
                    shares,                               # 持仓份额
                    mv if is_capital else mv_nav,         # 今日收盘市值
                    weight,                               # 今日收盘占资金权重
                    daily_pnl,                            # 当日损益
                    cum_pnl,                              # 持仓期间累计损益
                ])

            # --- 写入单日 Excel ---
            path = os.path.join(folder, f"{date_str}.xlsx")
            wb   = Workbook()
            ws   = wb.active
            ws.title = str(date)[:10]
            ws.row_dimensions[1].height = 22
            _write_header(ws, 1, HEADERS, WIDTHS, "capital")

            for row_idx, row_vals in enumerate(rows_data, start=2):
                for col_idx, val in enumerate(row_vals, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                _style_row(ws, row_idx, len(HEADERS), FMT_MAP)

            ws.freeze_panes = "C2"
            wb.save(path)

        print(f"  每日仓位明细 → {folder}  ({n_days} 个文件)")

    # ------------------------------------------------------------------ #
    # Sheet 5 - 净值序列（仅当有bench时）
    # ------------------------------------------------------------------ #

    def _write_nav_series_sheet(self, ws) -> None:
        """写入策略和基准的净值序列"""
        if self.result.bench_nav_series is None:
            ws.cell(1, 1, "无基准数据")
            return
        
        # 对齐两个净值序列
        strategy_nav = self.result.nav_series
        bench_nav = self.result.bench_nav_series
        
        # 取交集日期
        common_dates = strategy_nav.index.intersection(bench_nav.index)
        strategy_nav = strategy_nav.loc[common_dates]
        bench_nav = bench_nav.loc[common_dates]
        
        # 表头
        headers = ["日期", "策略净值", "基准净值", "相对强弱(策略/基准)"]
        widths = [14, 14, 14, 18]
        _write_header(ws, 1, headers, widths, "metrics")
        
        # 计算相对强弱
        relative_strength = strategy_nav / bench_nav
        
        # 数据行
        fmt_map = {2: "0.000000", 3: "0.000000", 4: "0.000000"}
        for i, date in enumerate(common_dates, start=2):
            ws.cell(i, 1, str(date)[:10])
            ws.cell(i, 2, float(strategy_nav.loc[date]))
            ws.cell(i, 3, float(bench_nav.loc[date]))
            ws.cell(i, 4, float(relative_strength.loc[date]))
            _style_row(ws, i, 4, fmt_map)
        
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Sheet 6 - 每日收益率（仅当有bench时）
    # ------------------------------------------------------------------ #

    def _write_daily_returns_sheet(self, ws) -> None:
        """写入策略和基准的每日收益率，以及超额收益"""
        if self.result.bench_nav_series is None:
            ws.cell(1, 1, "无基准数据")
            return
        
        # 对齐两个净值序列
        strategy_nav = self.result.nav_series
        bench_nav = self.result.bench_nav_series
        
        # 取交集日期
        common_dates = strategy_nav.index.intersection(bench_nav.index)
        strategy_nav = strategy_nav.loc[common_dates]
        bench_nav = bench_nav.loc[common_dates]
        
        # 计算日收益率
        strategy_returns = strategy_nav.pct_change().dropna()
        bench_returns = bench_nav.pct_change().dropna()
        
        # 对齐收益率日期
        common_ret_dates = strategy_returns.index.intersection(bench_returns.index)
        strategy_returns = strategy_returns.loc[common_ret_dates]
        bench_returns = bench_returns.loc[common_ret_dates]
        
        # 计算超额收益和累计超额
        excess_returns = strategy_returns - bench_returns
        cum_excess = (1 + excess_returns).cumprod()
        
        # 表头
        headers = ["日期", "策略收益率", "基准收益率", "超额收益", "累计超额收益"]
        widths = [14, 14, 14, 14, 16]
        _write_header(ws, 1, headers, widths, "metrics")
        
        # 数据行
        fmt_map = {2: "0.00%", 3: "0.00%", 4: "0.00%", 5: "0.000000"}
        for i, date in enumerate(common_ret_dates, start=2):
            ws.cell(i, 1, str(date)[:10])
            ws.cell(i, 2, float(strategy_returns.loc[date]))
            ws.cell(i, 3, float(bench_returns.loc[date]))
            ws.cell(i, 4, float(excess_returns.loc[date]))
            ws.cell(i, 5, float(cum_excess.loc[date]))
            _style_row(ws, i, 5, fmt_map)
        
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Sheet 7 - 月度收益（仅当有bench时）
    # ------------------------------------------------------------------ #

    def _write_monthly_returns_sheet(self, ws) -> None:
        """写入策略和基准的月度收益率，以及月度超额收益"""
        if self.result.bench_nav_series is None:
            ws.cell(1, 1, "无基准数据")
            return
        
        # 对齐两个净值序列
        strategy_nav = self.result.nav_series
        bench_nav = self.result.bench_nav_series
        
        # 取交集日期
        common_dates = strategy_nav.index.intersection(bench_nav.index)
        strategy_nav = strategy_nav.loc[common_dates]
        bench_nav = bench_nav.loc[common_dates]
        
        # 计算月度收益率
        strategy_monthly = strategy_nav.resample("ME").last().pct_change().dropna()
        bench_monthly = bench_nav.resample("ME").last().pct_change().dropna()
        
        # 对齐月度日期
        common_months = strategy_monthly.index.intersection(bench_monthly.index)
        strategy_monthly = strategy_monthly.loc[common_months]
        bench_monthly = bench_monthly.loc[common_months]
        
        # 计算月度超额收益
        excess_monthly = strategy_monthly - bench_monthly
        
        # 表头
        headers = ["月份", "策略月收益", "基准月收益", "超额月收益"]
        widths = [14, 14, 14, 14]
        _write_header(ws, 1, headers, widths, "metrics")
        
        # 数据行
        fmt_map = {2: "0.00%", 3: "0.00%", 4: "0.00%"}
        for i, month in enumerate(common_months, start=2):
            ws.cell(i, 1, str(month)[:7])  # YYYY-MM
            ws.cell(i, 2, float(strategy_monthly.loc[month]))
            ws.cell(i, 3, float(bench_monthly.loc[month]))
            ws.cell(i, 4, float(excess_monthly.loc[month]))
            _style_row(ws, i, 4, fmt_map)
        
        # 添加月度收益率色阶条件格式
        last_row = len(common_months) + 1
        if last_row > 1:
            for col in ["B", "C", "D"]:
                ws.conditional_formatting.add(
                    f"{col}2:{col}{last_row}",
                    ColorScaleRule(
                        start_type="min", start_color="F8696B",
                        mid_type="num", mid_value=0, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B",
                    ),
                )
        
        ws.freeze_panes = "A2"
